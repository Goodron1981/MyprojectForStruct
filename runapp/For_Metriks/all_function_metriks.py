import datetime
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, colors


def count_domens():
    # Считаем количество обработаных доменов
    wb = openpyxl.load_workbook('/home/max/websites/metriks.xlsx')
    sheet = wb['Metriks']
    last_row = len(sheet['A:A'])
    nowtime = datetime.datetime.now()
    formattime = nowtime.strftime('%d.%m.%Y %H:%M:%S')
    sheet['A' + str(last_row + 1)] = formattime
    wb.save('/home/max/websites/metriks.xlsx')

def count_articles():
    # Количество статей с фиксацией времени
    wb = openpyxl.load_workbook('/home/max/websites/metriks.xlsx')
    sheet = wb['Metriks']
    last_row = len(sheet['B:B'])
    nowtime = datetime.datetime.now()
    formattime = nowtime.strftime('%d.%m.%Y %H:%M:%S')
    sheet['B' + str(last_row + 1)] = formattime
    wb.save('/home/max/websites/metriks.xlsx')

def count_test_unikal():
    # Количество проверок на уникальность
    wb = openpyxl.load_workbook('/home/max/websites/metriks.xlsx')
    sheet = wb['Metriks']
    last_row = len(sheet['C:C'])
    nowtime = datetime.datetime.now()
    formattime = nowtime.strftime('%d.%m.%Y %H:%M:%S')
    sheet['C' + str(last_row + 1)] = formattime
    wb.save('/home/max/websites/metriks.xlsx')

def count_true_unikal():
    # Количество положительных проверок на уникальность
    wb = openpyxl.load_workbook('/home/max/websites/metriks.xlsx')
    sheet = wb['Metriks']
    last_row = len(sheet['D:D'])
    nowtime = datetime.datetime.now()
    formattime = nowtime.strftime('%d.%m.%Y %H:%M:%S')
    sheet['D' + str(last_row + 1)] = formattime
    wb.save('/home/max/websites/metriks.xlsx')

def mark_start():
    # Промаркировать начало обработки списка доменов
    wb = openpyxl.load_workbook('/home/max/websites/metriks.xlsx')
    sheet = wb['Metriks']
    last_row = len(sheet['A:A'])
    nowtime = datetime.datetime.now()
    formattime = nowtime.strftime('%d.%m.%Y %H:%M:%S')
    sheet['A' + str(last_row + 1)] = formattime
    my_red = openpyxl.styles.colors.Color(rgb='31c925')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    sheet['A' + str(last_row + 1)].fill = my_fill
    wb.save('/home/max/websites/metriks.xlsx')

def mark_fin():
    # Промаркировать окончание обработки списка доменов
    wb = openpyxl.load_workbook('/home/max/websites/metriks.xlsx')
    sheet = wb['Metriks']
    last_row = len(sheet['A:A'])
    nowtime = datetime.datetime.now()
    formattime = nowtime.strftime('%d.%m.%Y %H:%M:%S')
    sheet['A' + str(last_row + 1)] = formattime
    my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    sheet['A' + str(last_row + 1)].fill = my_fill
    wb.save('/home/max/websites/metriks.xlsx')

